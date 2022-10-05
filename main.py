import requests
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook, load_workbook

# Build for webmd
field = 'pediatrics' # change these to whatever you are searching for (make sure the final URL works!)
    # https://doctor.webmd.com/providers/specialty
state = 'California'
city = 'los-angeles'
End_Search = 0
page_number = 1
URL = 'https://doctor.webmd.com/providers/specialty/'+field+'/'+state+'/'+city
if city.find('-') != -1:
    shortened_city = city[0] + city[city.find('-') + 1]
else:
    shortened_city = city[0]


title_name = field + ' (' + shortened_city.upper() + ', ' + state[:2].upper() + ')'
wb = Workbook()             # open new workbook, use load_workbook if existing
ws = wb.active
ws.title = title_name
 
# write title row
ws.append(['Name', 'Specialization', 'Homepage','Telephone','Address'])

def search_url(field,state,city,page_num):
    URL = 'https://doctor.webmd.com/providers/specialty/'+field+'/'+state+'/'+city+'?pagenumber='+str(page_number)
    page = requests.get(URL)
    soup = BeautifulSoup(page.content, "html.parser")
    user_table = soup.find_all('div', class_='results-card-wrap')
    for user in user_table:
        user_links = (user.find_all('a')) # looks for all the a tags (happens to contain the user website, telephone num, and address)
        user_site = user_links[0].get('href')
        user_tele = user_links[len(user_links)-1].get('href')
        user_name = user.find('h2').get_text()
        user_spec = user.find('p', class_='prov-specialty').get_text()
        user_addr = user.find('span', class_='addr-text').get_text()
        # this part is for the sake of putting the data into a csv file
        # with open('out.txt', 'a') as f:
        #     print(user_name)
        #     print(user_name, file=f)
        #     print(user_spec, file=f)
        #     print(user_site, file=f)
        #     print(user_tele, file=f)
        #     print(user_addr, file=f)
        #     print('\n', file=f)
        #     print('----------------------------------------------------------', file=f) # separates the users
        #     print('\n', file=f)
        row = [user_name.encode('utf-8'), user_spec.encode('utf-8'), user_site.encode('utf-8'), user_tele.encode('utf-8'), user_addr.encode('utf-8')]  # construct a row: shown only for example purposes
        ws.append(row)          # could use ws.append(div) since each div is a list

file_name = title_name + '.xlsx'
while not End_Search:
    search_url(field,state,city, page_number)
    page_number+=1
    print(page_number)
    time.sleep(6)
    wb.save(file_name)                           # Save the file