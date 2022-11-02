import requests
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook, load_workbook

# Build for webmd
field = 'pediatrics' # change these to whatever you are searching for (make sure the final URL works!)
    # https://doctor.webmd.com/providers/specialty
state = 'California'
city = 'los-angeles'
End_Search = 0
page_number = 1
URL = 'https://doctor.webmd.com/providers/specialty/'+field+'/'+state+'/'+city
if city.find('-') != -1:
    shortened_city = city[0] + city[city.find('-') + 1]
else:
    shortened_city = city[0]


# title_name = "test" + field + ' (' + shortened_city.upper() + ', ' + state[:2].upper() + ')'
# wb = Workbook()             # open new workbook, use load_workbook if existing
# ws = wb.active
# ws.title = title_name
 
# write title row
# ws.append(['Name', 'Specialization', 'Homepage','Telephone','Address'])

def search_url(field,state,city,page_num):
    URL = 'https://doctor.webmd.com/providers/specialty/'+field+'/'+state+'/'+city+'?pagenumber='+str(page_number)
    page = requests.get(URL)
    soup = BeautifulSoup(page.content, "html.parser")
    user_table = soup.find_all('div', class_='results-card-wrap')
    for user in user_table:
        user_links = (user.find_all('a')) # looks for all the a tags (happens to contain the user website, telephone num, and address)
        user_site = user_links[0].get('href')
        user_tele = user_links[len(user_links)-1].get('href')
        user_name = user.find('h2').get_text()
        user_spec = user.find('p', class_='prov-specialty').get_text()
        user_addr = user.find('span', class_='addr-text').get_text()
        
        inner_search_term = user_name[:user_name.find(",")+1].lower()
        if (("ms. " in user_name.lower()) or ("mr. " in user_name.lower()) or ("dr. " in user_name.lower())):
            inner_search_term = inner_search_term[4:-1]
        elif (("ms." in user_name.lower()) or ("mr." in user_name.lower()) or ("dr." in user_name.lower())):
            inner_search_term = inner_search_term[3:-1]
        inner_search_term = inner_search_term.replace(" ", "-")
        inner_search_term = inner_search_term.replace(".", "")

        #special
        #inner_search_term = "vernon-arnold-cates"

        inner_URL = 'https://www.thefastfeed.com/doctors/' + inner_search_term
        inner_page = requests.get(inner_URL)
        inner_soup = BeautifulSoup(inner_page.content, "html.parser")
        try:
            links = inner_soup.find("div", class_="article-body")
            links = links.find_all("li")[3]
            for link in links:
                if ("-" in link.get_text()):
                    user_fax = link.get_text()
                    print(user_name, user_fax)
                    break
        except:
            continue

        # this part is for the sake of putting the data into a csv file
        # print(user_name)
        # print(user_spec)
        # print(user_site)
        # print(user_tele)
        # print(user_addr)
        # print('\n')
        # print('----------------------------------------------------------') # separates the users
        # print('\n')
        # ws.append(row)          # could use ws.append(div) since each div is a list

        #new
        # break

# file_name = title_name + '.xlsx'
while not End_Search:
    search_url(field,state,city, page_number)

    page_number+=1
    #if (page_number % 3 == 0):
    print(page_number)
# time.sleep(6)
# wb.save(file_name)                           # Save the file